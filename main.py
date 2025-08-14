import json
from utils.pdf_utils import extract_pages
from utils.parsing_utils import parse_toc_line, parse_sections
from utils.io_utils import save_jsonl, validate_json
from openpyxl import Workbook


def save_validation_report_xlsx(toc_valid, toc_errors, sec_valid, sec_errors, filepath):
    wb = Workbook()

    # --- TOC Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Category", "Valid Count", "Error Count"])
    ws_summary.append(["Table of Contents", toc_valid, len(toc_errors)])
    ws_summary.append(["Sections", sec_valid, len(sec_errors)])

    # --- TOC Errors Sheet ---
    ws_toc = wb.create_sheet(title="TOC Errors")
    ws_toc.append(["Line", "Error"])
    for err in toc_errors:
        ws_toc.append([err.get("line", ""), err.get("error", "")])

    # --- Section Errors Sheet ---
    ws_sec = wb.create_sheet(title="Section Errors")
    ws_sec.append(["Section ID", "Error"])
    for err in sec_errors:
        ws_sec.append([err.get("section_id", ""), err.get("error", "")])

    wb.save(filepath)


def main():
    pdf_path = "data/input.pdf"
    doc_title = "USB Power Delivery Specification"

    print("Extracting pages...")
    pages = extract_pages(pdf_path)

    # Load schemas
    with open("schemas/toc_schema.json") as f:
        toc_schema = json.load(f)
    with open("schemas/section_schema.json") as f:
        sec_schema = json.load(f)

    # ==== TOC Extraction & Validation ====
    toc_entries = []
    toc_validation_errors = []
    print("Parsing Table of Contents...")

    for page_num, page_text in enumerate(pages[:3], start=1):  # First 3 pages for TOC
        if not page_text:
            continue
        for line in page_text.splitlines():
            record = parse_toc_line(line, doc_title)
            if record:
                try:
                    validate_json(record, toc_schema)
                    toc_entries.append(record)
                except Exception as e:
                    toc_validation_errors.append({"line": line, "error": str(e)})

    if toc_entries:
        save_jsonl(toc_entries, "data/toc.jsonl")
    else:
        print("[WARNING] No Table of Contents entries found.")

    # ==== Section Extraction & Validation ====
    sections = parse_sections(pages, doc_title)
    section_validation_errors = []
    valid_sections = []

    print("Parsing sections...")
    for sec in sections:
        try:
            validate_json(sec, sec_schema)
            valid_sections.append(sec)
        except Exception as e:
            section_validation_errors.append({"section_id": sec.get("section_id"), "error": str(e)})

    if valid_sections:
        save_jsonl(valid_sections, "data/sections.jsonl")
    else:
        print("[WARNING] No valid sections found.")

    # ==== Excel Validation Report ====
    xlsx_path = "data/validation_report.xlsx"
    save_validation_report_xlsx(
        len(toc_entries), toc_validation_errors,
        len(valid_sections), section_validation_errors,
        xlsx_path
    )

    print(f"\n✅ Validation report saved to {xlsx_path}")


if __name__ == "__main__":
    main()
